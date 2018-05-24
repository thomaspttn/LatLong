"""
County Lat/Long Program (c) 2018 Thomas Patton
<tjpatton1@gmail.com>
github.com/thomaspttn

Automates entering lat/long into an excel sheet
"""
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
import time


# Start Selenium
driver = webdriver.Chrome('D:/Downloads/ChromeDriver/chromedriver_win32/chromedriver.exe')
driver.get("http://www.lat-long.com/Search.cfm?q=Yonkers&State=NY&County=&FeatureType=civil")


# Start Openpyxl
wb = load_workbook(filename='new_york.xlsx')
ws = wb.active


# This function submits the appropriate name into the search bar, and finds the appropriate link to the lat/long
def find_link(cell):
    city_name = cell.value
    # No Lat/Long for Sheriff reported crimes
    if not ("SHERIFF" in city_name):
        # Make a lat long search based on cell data and county name
        place_elem = driver.find_element_by_name("q")
        place_elem.clear()
        q = get_partial_link_text(city_name)
        place_elem.send_keys(q)
        county_elem = driver.find_element_by_name("County")
        #county_elem.send_keys(county_name)
        county_elem.submit()
        print(q)
        try:
            link_elem = driver.find_element_by_partial_link_text(q)
            link_elem.click()
            return True
        except NoSuchElementException:
            print("Not Found")
            # Enter "Not Found' into sheet
            b_cell = "B" + str(row_val)
            c_cell = "C" + str(row_val)
            ws[c_cell] = "n/a"
            ws[b_cell] = "n/a"
            driver.get("http://www.lat-long.com/Search.cfm?q=Yonkers&State=NY&County=&FeatureType=civil")
            return False
    else:
        sheriff_cell = "B" + str(row_val)
        ws[sheriff_cell] = "Not Found, Sheriff"


# Takes the latitude and longitude data off the web page and puts it into the appropriate excel cell
def get_lat_long():
    lat_long_data = driver.find_element_by_xpath(
        "/html/body/table/tbody/tr/td/font/table/tbody/tr/td[1]/table[1]/tbody/tr/td[1]"
        "/table/tbody/tr[2]/td/table/tbody/tr/td[1]/font[2]")
    print(parse_result(lat_long_data.text))
    cell_input = parse_result(lat_long_data.text)
    # Enter data into sheet
    b_cell = "B" + str(row_val)
    c_cell = "C" + str(row_val)
    d_cell = "D" + str(row_val)
    ws[b_cell] = float(cell_input[0])
    ws[c_cell] = float(cell_input[1])
    ws[d_cell] = "Computer Entered"
    driver.get("http://www.lat-long.com/Search.cfm?q=Yonkers&State=NY&County=&FeatureType=civil")


# Gets the appropriate searchable terms based off of
def get_partial_link_text(value):
    partial_link_text = ""
    # Turns the cell value into proper noun format
    for i in range(len(value)):
        if (i == 0) or (value[i-1] == " "):
            partial_link_text = partial_link_text + value[i].upper()
        else:
            partial_link_text = partial_link_text + value[i].lower()
    # Village and townships must be modified to get a proper search
    if ("Township" in value) or ("Village" in value) or ("Town" in value):
        j = 0
        shortened_text = ""
        while partial_link_text[j] != " ":
            shortened_text = shortened_text + partial_link_text[j]
            j = j + 1
        return shortened_text
    else:
        return partial_link_text


# Turns the web page lat/long data into just the decimal degrees as an array with length of 2, [lat, long]
def parse_result(string):
    dec_lat = []
    dec_long = []
    i = 0
    while string[i] != 'D':
        i = i + 1
    i = i + 27
    while string[i] != '\n':
        dec_lat.append(string[i])
        i = i + 1
    i = i + 12
    while i < len(string):
        dec_long.append(string[i])
        i = i + 1
    dec_lat = "".join(dec_lat)
    dec_long = "".join(dec_long)
    coordinates = [dec_lat, dec_long]
    return coordinates


# Keeps the name of the county updated for searching
def update_county(cell):
    if "COUNTY" in cell.value and "SHERIFF" not in cell.value:
        global county_name
        j = 0
        shortened_text = ""
        while cell.value[j] != " ":
            shortened_text = shortened_text + cell.value[j]
            j = j + 1
        county_name = shortened_text


# Closes the driver and saves the Excel sheet
def close():
    driver.close()
    wb.save('new_york.xlsx')


def main():
    print("Time Started: " + str(time.localtime()))
    start = time.time()
    time.sleep(1)
    global row_val
    row_val = 0
    for row in ws.iter_rows(min_row=0, max_col=1):
        row_val = row_val + 1
        for cell in row:
            update_county(cell)
            status = find_link(cell)
            if status:
                get_lat_long()
    close()
    end = time.time()
    print("Time Finished: " + str(time.localtime()))
    print("Time Elapsed: " + str(end - start))


if __name__ == "__main__":
    main()
