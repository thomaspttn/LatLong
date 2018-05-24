"""
County Lat/Long Program (c) 2018 Thomas Patton
<tjpatton1@gmail.com>
github.com/thomaspttn

Automates entering lat/long into an excel sheet
"""
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
import time

#Launch Function - Does most of program
#TODO: Separate into smaller functions
def launch():
    #Start Selenium
    driver = webdriver.Chrome('D:/Downloads/ChromeDriver/chromedriver_win32/chromedriver.exe')
    driver.get("http://www.lat-long.com/Ohio/")

    #Start Openpyxl
    wb = load_workbook(filename = 'latlong.xlsx')
    ws = wb.active

    #Initialize global vars
    countyName = ""
    rowVal = 0
    for row in ws.iter_rows(min_row=9, max_col=1, max_row=15):
        rowVal = rowVal + 1
        print(rowVal)
        for cell in row:
            cityName = cell.value
            #No Lat/Long for Sheriff reported crimes
            if not ("SHERIFF" in cityName):
                #Assign County Name
                if "COUNTY" in cityName:
                    countyName = removeEnd(cell.value)
                #Make a lat long search based on cell data and county name
                placeElem = driver.find_element_by_name("q")
                if "TOWNSHIP" in cityName:
                    placeElem.send_keys(removeEnd(cell.value))
                elif "VILLAGE" in cityName:
                    placeElem.send_keys(removeEnd(cell.value))
                else:
                    placeElem.send_keys(cell.value)
                countyElem = driver.find_element_by_name("County")
                countyElem.send_keys(countyName)
                countyElem.submit()
                properNameText = properName(cell.value)
                print(properNameText)
                cellInput = ""
                try:
                    if "TOWNSHIP" in cityName:
                        print("Searching for " + "'" + removeEnd(properName(cell.value)) + "'")
                        linkElem = driver.find_element_by_partial_link_text(properName(removeEnd(cell.value)))
                    elif "VILLAGE" in cityName:
                        linkElem = driver.find_element_by_partial_link_text(properName(removeEnd(cell.value)))
                    else:
                        linkElem = driver.find_element_by_partial_link_text(properNameText)
                    linkElem.click()
                    #Get lat/long values
                    latLong = driver.find_element_by_xpath("/html/body/table/tbody/tr/td/font/table/tbody/tr/td[1]/table[1]/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[1]/font[2]")
                    #print(latLong.text)
                    print(parseResult(latLong.text))
                    cellInput = parseResult(latLong.text)
                    #Enter data into sheet
                    bCell = "B" + str(rowVal)
                    ws[bCell] = float(cellInput[0])
                    cCell = "C" + str(rowVal)
                    ws[cCell] = float(cellInput[1])
                    driver.get("http://www.lat-long.com/Ohio/")
                except NoSuchElementException:
                    print("Not Found")
                    cellInput = "n/a"
                    #Enter "Not Found' into sheet
                    bCell = "B" + str(rowVal)
                    ws[bCell] = cellInput
                    cCell = "C" + str(rowVal)
                    ws[cCell] = cellInput
                    driver.get("http://www.lat-long.com/Ohio/")
                    break
            else:
                newCell = "B" + str(rowVal)
                ws[newCell] = "Not Found, Sheriff"
    #Post Entry, close everything down
    driver.close()
    wb.save('latlong.xlsx')



#Change caps names from spread into normal proper noun format
def properName(str):
    properNameList = []
    properNameList.append(str[0])
    i = 1
    if ' ' in str:
        while str[i] != ' ':
            properNameList.append(str[i].lower())
            i = i + 1
        properNameList.append(' ')
        properNameList.append(str[i + 1])
        i = i + 2
        while i < len(str):
            properNameList.append(str[i].lower())
            i = i + 1
        return "".join(properNameList)
    else:
        while i < len(str):
            properNameList.append(str[i].lower())
            i = i + 1
        return "".join(properNameList)

#String concatenation, get the part of county before "county"
def removeEnd(str):
    shortened = []
    i = 0
    while str[i] != ' ':
        shortened.append(str[i])
        i = i + 1
    return "".join(shortened)

def parseResult(str):
    decLat = []
    decLong = []
    i = 0;
    while str[i] != 'D':
        i = i + 1
    i = i + 27
    while str[i] != '\n':
        decLat.append(str[i])
        i = i + 1
    i = i + 12
    while i < len(str):
        decLong.append(str[i])
        i = i + 1
    decLat = "".join(decLat)
    decLong = "".join(decLong)
    coords = [decLat, decLong]
    return coords
#This will be changed in future to be divided into better subfunctions
def main():
    launch()

if __name__ == "__main__":
    main()